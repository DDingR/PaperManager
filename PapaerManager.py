#
#  PapaerManager
#  ver 0.1
#
#  crafted by DDing
# 

import openpyxl
from src.menu import *
from src.functions import *

def load_xlsx():
    xlsx_name = input('\n불러올 파일을 입력하세요: ')
    try:
        wb = openpyxl.load_workbook(xlsx_name)
    except FileNotFoundError:
        print('\033[31m' + 'ERR// 파일이 존재하지 않습니다' + '\033[0m')
        cmd = input('새로 만들겠습니까? (y/n): ')
        if cmd == 'y':
                xlsx_name = input('파일명을 입력하세요: ')
                wb = openpyxl.Workbook()
    except openpyxl.utils.exceptions.InvalidFileException:
        print('\033[31m' + 'ERR// 확장자가 잘못되었습니다' + '\033[0m')
        load_xlsx() # 여기서 에러 발생 # 밑의 return 문에서 발생하는듯

    return wb, xlsx_name

def load_data(ws):
    # 데이터의 크기를 구하기 위함
    for i in range(100): # 일단 100 으로 두었다
        if ws['A' + str(i + 1)].value == None:
            max_row = i
            break

    data = []
    for i in range(max_row):
        tmp = certainData()

        tmp.paper_name = ws['A' + str(i+1)].value
        tmp.keyWords = str(ws['B' + str(i+1)].value)
        tmp.author = ws['C' + str(i+1)].value
        tmp.date = ws['D' + str(i+1)].value 
        tmp.summary = ws['E' + str(i+1)].value
        tmp.link = ws['F' + str(i+1)].value
        data.append(tmp) 

    return data

def save_data(data, ws):
    for i in range(len(data)):
        ws['A' + str(i+1)] = data[i].paper_name
        ws['B' + str(i+1)] = data[i].keyWords
        ws['C' + str(i+1)] = data[i].author
        ws['D' + str(i+1)] = data[i].date
        ws['E' + str(i+1)] = data[i].summary
        ws['F' + str(i+1)] = data[i].link

if __name__ == "__main__":
    # load file
    wb, xlsx_name = load_xlsx()
    # ws = wb.get_sheet_by_name('main_sheet') # 해당 sheet 의 이름을 알아야 한다
    ws = wb.active # 활성화돼있는 sheet load (첫sheet)
    #  data 로드
    data = []
    data = load_data(ws)
    print('파일 및 데이터를 불러오기를 완료하였습니다\n')

    # 이제부터 실제 메뉴 display
    while(1): 
        display_menu()
        cmd = input('시작할 작업을 선택하세요: ')
        print('\n')

        if cmd == '1':
            print('=== 논문 검색하기 시작 ===')
            display_menu_search()
            cmd = input('무엇으로 검색하시겠습니까?: ')
            search(data, cmd)

        elif cmd == '2':
            print('=== 논문 수정하기 시작 ===')
            display_menu_search()
            cmd = input('무엇으로 수정하겠습니까?: ')
            # edit(data[i], cmd)
            print('\033[31m' + 'ERR// 아직 구현 중' + '\033[0m')

        elif cmd == '3':
            print('=== 저장된 전체 논문 조회 ===')
            for i in range(len(data)):
                display_certain_data(data[i], i)
            print('=== 전체 논문 조회 종료 ===\n')
            input('계속하려면 아무키나 입력하세요...\n')

        elif cmd == '4':
            print('새로운 논문 저장')
            tmp = certainData()
            tmp = register()
            data.append(tmp)

        elif cmd == '5':
            print('등록된 논문 삭제')
            cmd = input('삭제할 논문의 번호를 입력하세요: ')
            del data[int(cmd)-1]

        elif cmd == '6':
            print('논문 정렬하기')
            print('\033[31m' + 'ERR// 아직 구현 중' + '\033[0m')

        elif cmd == '0':
            print('작업을 종료합니다')

            cmd = input('저장할건가요? (y/n): ')
            if cmd == 'y':
                # data 저장
                save_data(data, ws)

                cmd = input('저장될 파일의 이름을 변경하겠습니까? (y/n): ')
                if cmd == 'y':
                    save_name = input('저장될 이름을 입력하세요 (.xlsx 포함): ')
                    wb.save(save_name)
                    print(save_name + ' 으로 저장되었습니다')
                else:
                    wb.save(xlsx_name)
                    print(xlsx_name + ' 으로 저장되었습니다')

                print('프로그램 종료' + '\n')

            exit()
        else:
            print('\033[31m' + 'ERR// 잘못된 명령어입니다' + '\033[0m')


            

